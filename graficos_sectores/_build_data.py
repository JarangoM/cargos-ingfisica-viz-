# Script de reconstrucción de sectores_datos.js a partir del Excel fuente.
# El Excel (BD_Graduados Ing_V2_actualizado.xlsx) no está en este repositorio
# por tamaño/privacidad; para regenerar los datos, copia este script a la
# carpeta local del proyecto (junto al Excel) y ejecútalo desde ahí.

import openpyxl, json, unicodedata
from collections import OrderedDict, defaultdict

wb = openpyxl.load_workbook('../BD_Graduados Ing_V2_actualizado.xlsx', read_only=True, data_only=True)

TYPO_FIX = {'investrigacion': 'investigacion'}


def strip_accents(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn')


def norm_key(s):
    s = s.strip().lower().replace('\xa0', ' ')
    s = ' '.join(s.split())
    s = TYPO_FIX.get(s, s)
    return strip_accents(s)


def parse_sheet(sheet_name):
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True))
    blocks = OrderedDict()
    col_starts = [0, 4, 8]
    current_header = {c: None for c in col_starts}
    for row in rows:
        for c in col_starts:
            name = row[c] if c < len(row) else None
            freq = row[c + 1] if c + 1 < len(row) else None
            if name is None and freq is None:
                continue
            if freq == 'Column1.Frecuencia':
                key = str(name).strip()
                blocks[key] = OrderedDict()
                current_header[c] = key
            elif name is not None and freq is not None:
                key = current_header[c]
                if key is None:
                    continue
                raw_label = str(name).strip().replace('\xa0', ' ')
                raw_label = ' '.join(raw_label.split())
                nk = norm_key(raw_label)
                bucket = blocks[key]
                if nk not in bucket:
                    bucket[nk] = {'frecuencia': 0}
                bucket[nk]['frecuencia'] += freq
    out = OrderedDict()
    for key, bucket in blocks.items():
        total = sum(v['frecuencia'] for v in bucket.values())
        items = []
        for nk, v in bucket.items():
            pct = round(v['frecuencia'] / total * 100, 2) if total else 0
            items.append({'key': nk, 'frecuencia': v['frecuencia'], 'porcentaje': pct})
        items.sort(key=lambda x: -x['frecuencia'])
        out[key] = items
    return out


nac = parse_sheet('graficos_sectores 2 nacional')
intl = parse_sheet('graficos_sectores 2 int.')

DISPLAY_NAMES = {
    'educacion': 'Educación',
    'tecnologia': 'Tecnología',
    'investigacion': 'Investigación',
    'consultoria': 'Consultoría',
    'finanzas': 'Finanzas',
    'industria': 'Industria',
    'mineria y energia': 'Minería y energía',
    'otros': 'Otros',
    'construccion': 'Construcción',
    'salud': 'Salud',
    'agricultura': 'Agricultura',
    'fabricacion de bebidas': 'Fabricación de bebidas',
    'servicios de alimentos y bebidas': 'Servicios de alimentos y bebidas',
    'transporte': 'Transporte',
    'servicios integrales de ingenieria': 'Servicios integrales de ingeniería',
    'fabricacion de papel': 'Fabricación de papel',
    'fabricacion de electrodomesticos y productos electricos y electronicos':
        'Fabricación de electrodomésticos y productos eléctricos y electrónicos',
}

COLORS_LIGHT = {
    'educacion': '#c80b0b',
    'tecnologia': '#00856d',
    'investigacion': '#c40a4c',
    'consultoria': '#089148',
    'finanzas': '#b90a85',
    'industria': '#078115',
    'mineria y energia': '#a909b4',
    'otros': '#238007',
    'construccion': '#8d0ce6',
    'salud': '#4b7a06',
    'agricultura': '#6538f5',
    'fabricacion de bebidas': '#6b7100',
    'servicios de alimentos y bebidas': '#384ff5',
    'transporte': '#8a6200',
    'servicios integrales de ingenieria': '#0b67ce',
    'fabricacion de papel': '#ae4409',
    'fabricacion de electrodomesticos y productos electricos y electronicos': '#0086a3',
}
COLORS_DARK = {
    'educacion': '#e20c0c',
    'tecnologia': '#008e75',
    'investigacion': '#dd0c55',
    'consultoria': '#049246',
    'finanzas': '#d10b96',
    'industria': '#089318',
    'mineria y energia': '#bf0bcb',
    'otros': '#289108',
    'construccion': '#a12af4',
    'salud': '#548b07',
    'agricultura': '#7852f6',
    'fabricacion de bebidas': '#798107',
    'servicios de alimentos y bebidas': '#5164f6',
    'transporte': '#9e6f00',
    'servicios integrales de ingenieria': '#0c74e9',
    'fabricacion de papel': '#c74d0a',
    'fabricacion de electrodomesticos y productos electricos y electronicos': '#0087a3',
}

all_keys = set()
for blocks in (nac, intl):
    for items in blocks.values():
        for it in items:
            all_keys.add(it['key'])
missing = [k for k in all_keys if k not in DISPLAY_NAMES or k not in COLORS_LIGHT or k not in COLORS_DARK]
print('Missing mapping for:', missing)
print('Total unique sectors:', len(all_keys))

PERIODOS = [
    ('2009_1_2014_2', '2009-1 a 2014-2'),
    ('2015_1_2019_2', '2015-1 a 2019-2'),
    ('2020_1_2021_2', '2020-1 a 2021-2'),
    ('2022_1_2025_2', '2022-1 a 2025-2'),
]


def build_dataset(blocks):
    periodos_out = []
    for pid, plabel in PERIODOS:
        trabajos = []
        for t in (1, 2, 3):
            block_key = 'sector{}_{}'.format(t, pid)
            items = blocks.get(block_key, [])
            items_out = [
                {
                    'sector': DISPLAY_NAMES[it['key']],
                    'categoria': it['key'],
                    'frecuencia': it['frecuencia'],
                    'porcentaje': it['porcentaje'],
                }
                for it in items
            ]
            total = sum(it['frecuencia'] for it in items_out)
            trabajos.append({'trabajo': t, 'total': total, 'items': items_out})
        periodos_out.append({'id': pid, 'label': plabel, 'trabajos': trabajos})
    return periodos_out


totals = defaultdict(int)
for blocks in (nac, intl):
    for items in blocks.values():
        for it in items:
            totals[it['key']] += it['frecuencia']
cat_order = sorted(totals.keys(), key=lambda k: -totals[k])

data = {
    'categorias': [
        {'id': k, 'nombre': DISPLAY_NAMES[k], 'colorClaro': COLORS_LIGHT[k], 'colorOscuro': COLORS_DARK[k]}
        for k in cat_order
    ],
    'nacional': build_dataset(nac),
    'internacional': build_dataset(intl),
}

js = 'const SECTORES_DATA = ' + json.dumps(data, ensure_ascii=False, indent=2) + ';\n'
with open('sectores_datos.js', 'w', encoding='utf-8') as f:
    f.write(js)
print('Wrote sectores_datos.js —', len(js), 'bytes,', len(cat_order), 'categorias')
